import numpy as np
import openpyxl
import pickle

for type_a in range(8):
    for type_b in range(1, 4):
        print(f'{type_a}_{type_b}')
        workbook = openpyxl.load_workbook('data/AUV/Temp/C' + f'{type_a}_{type_b}' + '.xlsx')
        sheet_names = workbook.sheetnames
        sheet1 = workbook[sheet_names[0]]

        #统计行数和列数
        rows = sheet1.max_row
        cols = sheet1.max_column

        #读出数据
        datas = []
        for row in sheet1.rows:
            data = []
            for cell in row:
                data.append(cell.value)
            datas.append(data)

        # 提取每隔300个二级列表中的1024个，组成新的一组列表
        new_lists = []
        flattened_lists = []
        for i in range(0, len(datas), 300):
            if i + 1024 <= len(datas):
                new_list = datas[i:i + 1024]
                new_lists.append(new_list)
                flattened_list = []
                for i in range(3):
                    for sublist in new_list:
                        flattened_list.append(sublist[i])
                flattened_lists.append(flattened_list)

        # 将flattened_lists中的数据输出
        output = np.array(flattened_lists)
        file = open('./data/AUV/C' + f'{type_a}_{type_b}' + '.pkl', 'wb')
        pickle.dump(output, file)
        file.close()


